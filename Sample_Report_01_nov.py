import os
import pandas as pd
import openpyxl
import os.path
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
import numpy as np
import collections

def csvFileRead():
    try: 
       
        if os.path.isfile('workbook1.xlsx'):
            
        #----------------------groups--------------
            # Reading the csv file
            df = pd.read_csv("groups.csv", sep=',')
            #print(df)
            
            #remove duplicates
            df_group= df.drop_duplicates()
            #print(df_group)

            #gropsws sheet data
            df_name=pd.DataFrame({'name':df_group['name']})
            df_name1=df_name.drop_duplicates().dropna()
            #print(df_name1)
            
            #bycluster sheet data
            name=list(df_name1['name'])
            #print(len(name))

            #find environment values
            env=[]
            for i in range(len(name)):
                if 'PROD' in name[i] or 'prod' in name[i] or 'Prod' in name[i]:
                    a='PROD'
                    env.append(a)
                elif 'TPT' in name[i] or 'tpt' in name[i]:
                    b='TPT'
                    env.append(b)
                else:
                    c='NON-PROD'
                    env.append(c)       

            df_bycluster=pd.DataFrame({'Name':name,'Environment':env})
            #print(df_bycluster)
            
        #----------------------databases--------------------------
            # Reading the csv file
            df_db = pd.read_csv("databases.csv", sep=',')
            #print(df_db)
            
            #remove duplicates
            df_db1= df_db.drop_duplicates().dropna()
            #print(df_db1.columns)

            print('Add column name of database extract sheet:')
            col=input('Please enter comma separated 9 column names like name, version etc. :')
            #number,name,id1,id2,hostname,port,typename,version,database
            header_list=col.split(',')
            df_db1.columns=header_list
            #print(df_db1)

            #databaseWS creation
            df_dbws=pd.DataFrame({'name':list(df_db1[header_list[1]]),'version':list(df_db1[header_list[7]]),'database':list(df_db1[header_list[8]])})
            df_dbws1=df_dbws.drop_duplicates()
            df_dbws11 = df_dbws1[~df_dbws1[header_list[8]].isin(['config','local'])]
            df_dbws_final = df_dbws11[~df_dbws11[header_list[1]].isin(['admin'])]
            #print(df_dbws_final)


            #find environment values
            db_name=list(df_dbws_final['name'])
            db_env=[]
            for i in range(len(db_name)):
                if 'PROD' in db_name[i] or 'prod' in db_name[i] or 'Prod' in db_name[i]:
                    if 'NON-PROD' in db_name[i] or 'non-prod' in db_name[i] or 'Non-Prod' in db_name[i]:
                        a='NON-PROD'
                        db_env.append(a)
                    else:
                        a='PROD'
                        db_env.append(a)
                elif 'TPT' in db_name[i] or 'tpt' in db_name[i]:
                    b='TPT'
                    db_env.append(b)
                else:
                    c='NON-PROD'
                    db_env.append(c) 

            #bydatabase creation
            df_bdb=pd.DataFrame({'NAME':list(df_dbws_final['name']),'ENVIRONMENT': db_env,'MVERSION':list(df_dbws_final['version']),'DATABASENAME':list(df_dbws_final['database'])})
            #print(df_bdb)

        #--------------------hosts--------------------------------
            # Reading the csv file
            df_hostE = pd.read_csv("host.csv", sep=',')
            df_host=df_hostE.dropna()
            #print(df_host)

            #serverws creation
            df_host=pd.DataFrame({'name':list(df_host['name']),'hostname':list(df_host['hostname']),'version':list(df_host['version'])})
            df_host1=df_host.drop_duplicates()
            #print(df_host1)
            df_hostws = df_host1[~df_host1['name'].isin(['admin'])]
            #print(df_hostws)

            #find environment values
            h_name=list(df_hostws['name'])
            host_env=[]
            for i in range(len(h_name)):
                if 'PROD' in h_name[i] or 'prod' in h_name[i] or 'Prod' in h_name[i]:
                    if 'NON-PROD' in h_name[i] or 'non-prod' in h_name[i] or 'Non-Prod' in h_name[i]:
                        a='NON-PROD'
                        host_env.append(a)
                    else:
                        a='PROD'
                        host_env.append(a)
                elif 'TPT' in h_name[i] or 'tpt' in h_name[i]:
                    b='TPT'
                    host_env.append(b)
                else:
                    c='NON-PROD'
                    host_env.append(c) 

            host_L=list(df_hostws['name'])
            #print(host_L)
            hostn_L=list(df_hostws['hostname'])
            #print(hostn_L)

            #Byserver creation
            df_hostbs=pd.DataFrame({'Name':host_L,'Hostname':hostn_L,'Environment':host_env})
            #print(df_hostbs)


            host_L1=list(df_hostE['name'])
            #print(host_L1)
            hostn_L1=list(df_hostE['hostname'])
            #print(hostn_L1)
            
            host_tn_L=list(df_hostE['typeName'])
            #print(host_tn_L)
                         
            #DatanodeWS creation
            df_dnws=pd.DataFrame({'Name':host_L1,'Hostname':hostn_L1,'Typename':host_tn_L})
            #print(df_dnws)
            df_dnws11=df_dnws[df_dnws['Typename'].isin(['REPLICA_PRIMARY','REPLICA_SECONDARY'])]
            df_dnws1=df_dnws11.drop_duplicates()
            #print(df_dnws1)

            #find environment values
            h_name1=list(df_dnws1['Name'])
            h_hostname1=list(df_dnws1['Hostname'])
            host_env1=[]
            for i in range(len(h_name1)):
                if 'PROD' in h_name1[i] or 'prod' in h_name1[i] or 'Prod' in h_name1[i]:
                    if 'NON-PROD' in h_name1[i] or 'non-prod' in h_name1[i] or 'Non-Prod' in h_name1[i]:
                        a='NON-PROD'
                        host_env1.append(a)
                    else:
                        a='PROD'
                        host_env1.append(a)
                elif 'TPT' in h_name1[i] or 'tpt' in h_name1[i]:
                    b='TPT'
                    host_env1.append(b)
                else:
                    c='NON-PROD'
                    host_env1.append(c)
                    
            #Bydatanode creation
            df_bdn=pd.DataFrame({'Name':h_name1,'Hostname':h_hostname1,'Environment':host_env1})
            #print(df_bdn)

            #----------------------------------All sheets other than Growth, Summary and Repports sheet
           
            with pd.ExcelWriter('workbook1.xlsx',engine='openpyxl',mode='a') as writer:
                df_group.to_excel(writer, sheet_name='Groups Extract',index=False)
                df_name1.to_excel(writer, sheet_name='GroupsWS',index=False)
                df_bycluster.to_excel(writer, sheet_name='ByCluster',index=False)

                df_db1.to_excel(writer, sheet_name='Database Extract',index=False, header=False)
                df_dbws_final.to_excel(writer, sheet_name='DatabaseWS',index=False)
                df_bdb.to_excel(writer, sheet_name='ByDatabase',index=False)

                df_host.to_excel(writer, sheet_name='Hosts Extract',index=False)
                df_hostws.to_excel(writer, sheet_name='ServerWS',index=False)
                df_hostbs.to_excel(writer, sheet_name='ByServer',index=False)
                df_dnws1.to_excel(writer, sheet_name='DatanodeWS',index=False)
                df_bdn.to_excel(writer, sheet_name='ByDataNode',index=False)
                



          #--------------------Reports----------------------------
            wb = load_workbook('workbook1.xlsx')
            ws1 = wb.create_sheet('sheet',0)
            ws1.title='Reports'
            wb.save(filename = 'workbook1.xlsx')
            wb = load_workbook('workbook1.xlsx')
            sheet = wb['Reports']

            #Prod,Non-prod and TPT count calculation---------------------
            Server_count=dict(df_hostbs['Environment'].value_counts())
            #print(Server_count)
                
            #check dictionary has data for 3 environments or not
            if len(Server_count)!=3:
                if 'PROD' not in Server_count:
                    Server_count['PROD']=0

                if 'NON-PROD' not in Server_count:
                    Server_count['NON-PROD']=0
            
                if 'TPT' not in Server_count:
                    Server_count['TPT']=0

            cluster_count=dict(df_bycluster['Environment'].value_counts())
            #print(cluster_count)

            #check dictionary has data for 3 environments or not
            if len(cluster_count)!=3:
                if 'PROD' not in cluster_count:
                    cluster_count['PROD']=0

                if 'NON-PROD' not in cluster_count:
                    cluster_count['NON-PROD']=0
            

                if 'TPT' not in cluster_count:
                    cluster_count['TPT']=0

            database_count=dict(df_bdb['ENVIRONMENT'].value_counts())
            #print(database_count)

            #check dictionary has data for 3 environments or not
            if len(database_count)!=3:
                if 'PROD' not in database_count:
                    database_count['PROD']=0

                if 'NON-PROD' not in database_count:
                    database_count['NON-PROD']=0
            
                if 'TPT' not in database_count:
                    database_count['TPT']=0
    
            datanode_count=dict(df_bdn['Environment'].value_counts())
            #print(datanode_count)

            #check dictionary has data for 3 environments or not
            if len(datanode_count)!=3:
                if 'PROD' not in datanode_count:
                    datanode_count['PROD']=0

                if 'NON-PROD' not in datanode_count:
                    datanode_count['NON-PROD']=0
            
                if 'TPT' not in datanode_count:
                    datanode_count['TPT']=0

            #header of report sheet
            rows_1=[['Servers Count','','','','','Databases Count','','','','','Clusters Count','','',
                   'Data Nodes Count(Licensing)','','',''],
                  ['Environment','Name','Hostname','Total','','Environment','Name','Database Name','Total','',
                   'Row labels','Count of Name','','Environment','Name','Hostname','Total'],
                  ['PROD','','',Server_count['PROD'],'','PROD','','',database_count['PROD'],'',
                   'PROD',cluster_count['PROD'],'','PROD','','',datanode_count['PROD']]
                 ]

            for i in rows_1:
               sheet.append(i)


            #-----------------server table

            name_s=list(df_hostbs['Name'])
            host_s=list(df_hostbs['Hostname'])
            env_s=list(df_hostbs['Environment'])
            #print(name_s)
            #print(host_s)
            #print(env_s)

            n1,n2,n3=[],[],[]
            h1,h2,h3=[],[],[]

            for i in range(len(name_s)):
                if env_s[i] == 'PROD':
                   n1.append(name_s[i])
                   h1.append(host_s[i])

                elif env_s[i]=='NON-PROD':
                    n2.append(name_s[i])
                    h2.append(host_s[i])

                elif env_s[i]=='TPT':
                    n3.append(name_s[i])
                    h3.append(host_s[i])

            #print(n1)
            #print(n2)
            #print(n3)

            #---------------prod data---------
            x1 = np.array(n1)
            u1=np.unique(x1)
            
            dict1= collections.defaultdict(list)
            for i in range(len(u1)):
                #print('u1[i]',u1[i])
               
                for j in range(len(n1)):
                   # print('n1[j]',n1[j])
                   
                    if n1[j]==u1[i]:
                       # print('equal')
                        dict1[u1[i]].append(h1[j])

            p_name1=dict(dict1)
            #print('pn',p_name1)
                
                
            #------------------------non-prod data
           
            x2 = np.array(n2)
            u2=np.unique(x2)
            
            dict2= collections.defaultdict(list)
            for i in range(len(u2)):
                #print('u2[i]',u2[i])
               
                for j in range(len(n2)):
                    #print('n2[j]',n2[j])
                   
                    if n2[j]==u2[i]:
                        #print('equal')
                        dict2[u2[i]].append(h2[j])

            np_name2=dict(dict2)
            #print(np_name2)

            #------------------------TPT data
            x3 = np.array(n3)
            u3=np.unique(x3)
            #print(u3)
            
            dict3= collections.defaultdict(list)
            for i in range(len(u3)):
                #print('u3[i]',u3[i])
                for j in range(len(n3)):
                    #print('n3[j]',n3[j])
                   
                    if n3[j]==u3[i]:
                       # print('equal')
                        dict3[u3[i]].append(h3[j])

            t_name3=dict(dict3)
            #print(t_name3)             

            #-------------server count table design
            p_k=list(p_name1.keys())
            
            for i in p_k:
               row1=['',i,'','']
               sheet.append(row1)
               p_v=p_name1[i]
               for j in range(len(p_v)):
                   row2=['','',p_v[j],1]
                   sheet.append(row2)
    
               sheet.append(['',str(i+' Sub Total'),'',len(p_v)])

           
            row_2=['NON-PROD','','',Server_count['NON-PROD']]
            sheet.append(row_2)

            np_k=list(np_name2.keys())
            
            for i in np_k:
               row11=['',i,'','']
               sheet.append(row11)
               np_v=np_name2[i]
               for j in range(len(np_v)):
                   row21=['','',np_v[j],1]
                   sheet.append(row21)
    
               sheet.append(['',str(i+' Sub Total'),'',len(np_v)])

           
            row_3=['TPT','','',Server_count['TPT']]
            sheet.append(row_3)
            
            t_k=list(t_name3.keys())

            for i in t_k:
               row12=['',i,'','']
               sheet.append(row12)
               t_v=t_name3[i]
               for j in range(len(t_v)):
                   row22=['','',t_v[j],1]
                   sheet.append(row22)
    
               sheet.append(['',str(i+' Sub Total'),'',len(t_v)])

            sheet.append(['Grand Total','','',(Server_count['TPT']+Server_count['PROD']+Server_count['NON-PROD'])])

            #databases count table --------------------------------
            
            name_s1=list(df_bdb['NAME'])
            host_s1=list(df_bdb['DATABASENAME'])
            env_s1=list(df_bdb['ENVIRONMENT'])

            n11,n21,n31=[],[],[]
            h11,h21,h31=[],[],[]

            for i in range(len(name_s1)):
                if env_s1[i] == 'PROD':
                   n11.append(name_s1[i])
                   h11.append(host_s1[i])

                elif env_s1[i]=='NON-PROD':
                    n21.append(name_s1[i])
                    h21.append(host_s1[i])

                elif env_s1[i]=='TPT':
                    n31.append(name_s1[i])
                    h31.append(host_s1[i])

            #print('n11',n11,h11)
            #print('n21',n21,h21)
            #print('n31',n31,h31)
            

            #---------------prod data---------
            x11 = np.array(n11)
            u11=np.unique(x11)
            
            dict11= collections.defaultdict(list)
            for i in range(len(u11)):
                for j in range(len(n11)):
                    if n11[j]==u11[i]:
                        dict11[u11[i]].append(h11[j])

            p_name11=dict(dict11)
            #print('pn',p_name11)
                
                
            #------------------------non-prod data
           
            x21 = np.array(n21)
            u21=np.unique(x21)
            
            dict21= collections.defaultdict(list)
            for i in range(len(u21)):
               
                for j in range(len(n21)):
                   
                    if n21[j]==u21[i]:
                        dict21[u21[i]].append(h21[j])

            np_name21=dict(dict21)
            #print(np_name21)

            #------------------------TPT data
            x31 = np.array(n31)
            u31=np.unique(x31)
            #print(u31)
            
            dict31= collections.defaultdict(list)
            for i in range(len(u31)):
                for j in range(len(n31)):
                   
                    if n31[j]==u31[i]:
                        dict31[u31[i]].append(h31[j])

            t_name31=dict(dict31)
            #print(t_name31)


            #-------------database count table design

            if database_count['PROD']==0:
                k=4
                sheet['F4']='NON-PROD'
                sheet['I4']=database_count['NON-PROD']
                if database_count['NON-PROD']==0:
                    sheet['F4']='NON-PROD'
                    sheet['I4']=0
                    sheet['F5']='TPT'
                    sheet['I5']=database_count['TPT']
                    k=6
                    if database_count['TPT']==0:
                        sheet['F5']='TPT'
                        sheet['I5']=0
                        sheet['F6']='Grand Total'
                        sheet['I6']=database_count['TPT']+database_count['PROD']+database_count['NON-PROD']

                    else:
                        sheet['F5']='TPT'
                        sheet['I5']=database_count['TPT']
                    
                        t_k1=list(t_name31.keys())

                        for i in t_k1:
                            sheet[str('G'+str(k))]=str(i)
                            t_v1=t_name31[i]
                            for j in range(len(t_v1)):
                                sheet[str('H'+str(k+1+j))]=str(t_v1[j])
                                sheet[str('I'+str(k+1+j))]=1
                            sheet[str('F'+str(k+len(t_v1)+1))]=str(i +' SubTotal')
                            sheet[str('I'+str(k+len(t_v1)+1))]=len(t_v1)
                            k=k+len(t_v1)+2

                        sheet[str('F'+str(k))]='Grand Total'
                        sheet[str('I'+str(k))]=database_count['TPT']+database_count['PROD']+database_count['NON-PROD']


                else:
                    k=4
                    sheet[str('F'+str(k))]='NON-PROD'
                    sheet[str('I'+str(k))]=database_count['NON-PROD']

                    np_k1=list(np_name21.keys())
                    
                    for i in np_k1:
                        sheet[str('G'+str(k))]=str(i)
                        np_v1=np_name21[i]
                        for j in range(len(np_v1)):
                            sheet[str('H'+str(k+1+j))]=str(np_v1[j])
                            sheet[str('I'+str(k+1+j))]=1
                        sheet[str('F'+str(k+len(np_v1)+1))]=str(i +' SubTotal')
                        sheet[str('I'+str(k+len(np_v1)+1))]=len(np_v1)
                        k=k+len(np_v1)+2

                    if database_count['TPT']==0:
                        sheet[str('F'+str(k))]='TPT'
                        sheet[str('I'+str(k))]=0
                        sheet[str('F'+str(k+1))]='Grand Total'
                        sheet[str('I'+str(k+1))]=database_count['TPT']+database_count['PROD']+database_count['NON-PROD']

                    else:
                        sheet[str('F'+str(k))]='TPT'
                        sheet[str('I'+str(k))]=database_count['TPT']
                    
                        t_k1=list(t_name31.keys())

                        for i in t_k1:
                            sheet[str('G'+str(k))]=str(i)
                            t_v1=t_name31[i]
                            for j in range(len(t_v1)):
                                sheet[str('H'+str(k+1+j))]=str(t_v1[j])
                                sheet[str('I'+str(k+1+j))]=1
                            sheet[str('F'+str(k+len(t_v1)+1))]=str(i +' SubTotal')
                            sheet[str('I'+str(k+len(t_v1)+1))]=len(t_v1)
                            k=k+len(t_v1)+2

                        sheet[str('F'+str(k))]='Grand Total'
                        sheet[str('I'+str(k))]=database_count['TPT']+database_count['PROD']+database_count['NON-PROD']

            else:
                p_k1=list(p_name11.keys())
                #print(p_k1)

                k=4
                for i in p_k1:
                    #print('i',i)
                    #print('G'+str(k))
                    sheet[str('G'+str(k))]=str(i)
                    p_v1=p_name11[i]
                    #print(p_v1)
                    for j in range(len(p_v1)):
                        #print('j',j)
                        #print('H'+str(k+1+j))
                        sheet[str('H'+str(k+1+j))]=str(p_v1[j])
                        #print(str('I'+str(k+1+j)))
                        sheet[str('I'+str(k+1+j))]=1
                    sheet[str('F'+str(k+len(p_v1)+1))]=str(i +' SubTotal')
                    sheet[str('I'+str(k+len(p_v1)+1))]=len(p_v1)
                    k=k+len(p_v1)+2
                    #print(k)

                if database_count['NON-PROD']==0:
                    sheet[str('F'+str(k))]='NON-PROD'
                    sheet[str('I'+str(k))]=0
                    sheet[str('F'+str(k+1))]='TPT'
                    sheet[str('I'+str(k+1))]=database_count['TPT']
                    if database_count['TPT']==0:
                        sheet[str('F'+str(k))]='TPT'
                        sheet[str('I'+str(k))]=0
                        #print('F'+str(k+1))
                        sheet[str('F'+str(k+1))]='Grand Total'
                        sheet[str('I'+str(k+1))]=database_count['TPT']+database_count['PROD']+database_count['NON-PROD']

                    else:
                        sheet[str('F'+str(k))]='TPT'
                        sheet[str('I'+str(k))]=database_count['TPT']
                    
                        t_k1=list(t_name31.keys())

                        for i in t_k1:
                            sheet[str('G'+str(k))]=str(i)
                            t_v1=t_name31[i]
                            for j in range(len(t_v1)):
                                sheet[str('H'+str(k+1+j))]=str(t_v1[j])
                                sheet[str('I'+str(k+1+j))]=1
                            sheet[str('F'+str(k+len(t_v1)+1))]=str(i +' SubTotal')
                            sheet[str('I'+str(k+len(t_v1)+1))]=len(t_v1)
                            k=k+len(t_v1)+2

                        sheet[str('F'+str(k))]='Grand Total'
                        sheet[str('I'+str(k))]=database_count['TPT']+database_count['PROD']+database_count['NON-PROD']


                else:
                    sheet[str('F'+str(k))]='NON-PROD'
                    sheet[str('I'+str(k))]=database_count['NON-PROD']

                    np_k1=list(np_name21.keys())
                    
                    for i in np_k1:
                        sheet[str('G'+str(k))]=str(i)
                        np_v1=np_name21[i]
                        for j in range(len(np_v1)):
                            sheet[str('H'+str(k+1+j))]=str(np_v1[j])
                            sheet[str('I'+str(k+1+j))]=1
                        sheet[str('F'+str(k+len(np_v1)+1))]=str(i +' SubTotal')
                        sheet[str('I'+str(k+len(np_v1)+1))]=len(np_v1)
                        k=k+len(np_v1)+2

                    if database_count['TPT']==0:
                        sheet[str('F'+str(k))]='TPT'
                        sheet[str('I'+str(k))]=0
                        #print('F'+str(k+1))
                        sheet[str('F'+str(k+1))]='Grand Total'
                        sheet[str('I'+str(k+1))]=database_count['TPT']+database_count['PROD']+database_count['NON-PROD']

                    else:
                        sheet[str('F'+str(k))]='TPT'
                        sheet[str('I'+str(k))]=database_count['TPT']
                    
                        t_k1=list(t_name31.keys())

                        for i in t_k1:
                            sheet[str('G'+str(k))]=str(i)
                            t_v1=t_name31[i]
                            for j in range(len(t_v1)):
                                sheet[str('H'+str(k+1+j))]=str(t_v1[j])
                                sheet[str('I'+str(k+1+j))]=1
                            sheet[str('F'+str(k+len(t_v1)+1))]=str(i +' SubTotal')
                            sheet[str('I'+str(k+len(t_v1)+1))]=len(t_v1)
                            k=k+len(t_v1)+2

                        sheet[str('F'+str(k))]='Grand Total'
                        sheet[str('I'+str(k))]=database_count['TPT']+database_count['PROD']+database_count['NON-PROD']

            #cluster table----------------------------


            name_s2=list(df_bycluster['Name'])
            env_s2=list(df_bycluster['Environment'])
            #print(name_s2)
            #print(env_s2)
            #print(len(name_s2))

            n12,n22,n32=[],[],[]

            for i in range(len(name_s2)):
                if env_s2[i] == 'PROD':
                   n12.append(name_s2[i])
                   #print(name_s2[i])

                elif env_s2[i]=='NON-PROD':
                    n22.append(name_s2[i])
                    #print(name_s2[i])

                elif env_s2[i]=='TPT':
                    n32.append(name_s2[i])
                    #print(name_s2[i])

            #print(n12)
            #print(n22)
            #print(n32)

            #-------------cluster count table design

            if cluster_count['PROD']==0:
                k=4
                sheet['K4']='NON-PROD'
                sheet['L4']=cluster_count['NON-PROD']
                if cluster_count['NON-PROD']==0:
                    sheet['K4']='NON-PROD'
                    sheet['L4']=0
                    sheet['K5']='TPT'
                    sheet['L5']=cluster_count['TPT']
                    
                    if cluster_count['TPT']==0:
                        sheet['K5']='TPT'
                        sheet['L5']=0
                        sheet['K6']='Grand Total'
                        sheet['L6']=cluster_count['TPT']+cluster_count['PROD']+cluster_count['NON-PROD']

                    else:
                        k=5
                        sheet['K5']='TPT'
                        sheet['L5']=cluster_count['TPT']
                        for j in range(len(n32)):
                            sheet[str('K'+str(k+1+j))]=str(n32[j])
                            sheet[str('L'+str(k+1+j))]=1
                        k=k+len(n32)+1
                        sheet[str('K'+str(k))]='Grand Total'
                        sheet[str('L'+str(k))]=cluster_count['TPT']+cluster_count['PROD']+cluster_count['NON-PROD']


                else:
                    k=4
                    sheet[str('K'+str(k))]='NON-PROD'
                    sheet[str('L'+str(k))]=cluster_count['NON-PROD']

                    
                    for j in range(len(n22)):
                        sheet[str('K'+str(k+1+j))]=str(n22[j])
                        sheet[str('L'+str(k+1+j))]=1
                        
                    k=k+len(n22)+1
                    
                    if cluster_count['TPT']==0:
                        sheet['K'+str(k+1)]='TPT'
                        sheet['L'+str(k+1)]=0
                        sheet['K'+str(k+2)]='Grand Total'
                        sheet['L'+str(k+2)]=cluster_count['TPT']+cluster_count['PROD']+cluster_count['NON-PROD']

                    else:
                        sheet['K'+str(k+1)]='TPT'
                        sheet['L'+str(k+1)]=cluster_count['TPT']
                        for j in range(len(n32)):
                            sheet[str('K'+str(k+2+j))]=str(n32[j])
                            sheet[str('L'+str(k+2+j))]=1
                        k=k+len(n32)+1
                        sheet[str('K'+str(k))]='Grand Total'
                        sheet[str('L'+str(k))]=cluster_count['TPT']+cluster_count['PROD']+cluster_count['NON-PROD']

            else:

                k=4
                
                for j in range(len(n12)):
                    sheet[str('K'+str(k+1+j))]=str(n12[j])
                    sheet[str('L'+str(k+1+j))]=1
                    
                k=k+len(n12)+1
                if cluster_count['NON-PROD']==0:
                    sheet[str('K'+str(k))]='NON-PROD'
                    sheet[str('L'+str(k))]=0
                    sheet[str('K'+str(k+1))]='TPT'
                    sheet[str('L'+str(k+1))]=cluster_count['TPT']
                    if cluster_count['TPT']==0:
                        sheet[str('K'+str(k+1))]='TPT'
                        sheet[str('L'+str(k+1))]=0
                        
                        sheet[str('K'+str(k+2))]='Grand Total'
                        sheet[str('L'+str(k+2))]=cluster_count['TPT']+cluster_count['PROD']+cluster_count['NON-PROD']

                    else:
                        sheet[str('K'+str(k+1))]='TPT'
                        sheet[str('L'+str(k+1))]=cluster_count['TPT']
                    
                        for j in range(len(n32)):
                            sheet[str('K'+str(k+2+j))]=str(n32[j])
                            sheet[str('L'+str(k+2+j))]=1
                          
                        k=k+len(n32)+1
                        sheet[str('K'+str(k))]='Grand Total'
                        sheet[str('L'+str(k))]=cluster_count['TPT']+cluster_count['PROD']+cluster_count['NON-PROD']


                else:
                    sheet[str('K'+str(k))]='NON-PROD'
                    sheet[str('L'+str(k))]=cluster_count['NON-PROD']
                    for j in range(len(n22)):
                        sheet[str('K'+str(k+1+j))]=str(n22[j])
                        sheet[str('L'+str(k+1+j))]=1

                    k=k+len(n22)+1 
                    if cluster_count['TPT']==0:
                        sheet[str('K'+str(k))]='TPT'
                        sheet[str('L'+str(k))]=0
                    
                        sheet[str('K'+str(k+1))]='Grand Total'
                        sheet[str('L'+str(k+1))]=cluster_count['TPT']+cluster_count['PROD']+cluster_count['NON-PROD']

                    else:
                        sheet[str('K'+str(k))]='TPT'
                        sheet[str('L'+str(k))]=cluster_count['TPT']
                        for j in range(len(n32)):
                            sheet[str('K'+str(k+1+j))]=str(n32[j])
                            sheet[str('L'+str(k+1+j))]=1
                        k=k+len(n32)+1   
                        sheet[str('K'+str(k))]='Grand Total'
                        sheet[str('L'+str(k))]=cluster_count['TPT']+cluster_count['PROD']+cluster_count['NON-PROD']


            #datanode count table --------------------------------
            
            name_s1=list(df_bdb['NAME'])
            host_s1=list(df_bdb['DATABASENAME'])
            env_s1=list(df_bdb['ENVIRONMENT'])

            n11,n21,n31=[],[],[]
            h11,h21,h31=[],[],[]

            for i in range(len(name_s1)):
                if env_s1[i] == 'PROD':
                   n11.append(name_s1[i])
                   h11.append(host_s1[i])

                elif env_s1[i]=='NON-PROD':
                    n21.append(name_s1[i])
                    h21.append(host_s1[i])

                elif env_s1[i]=='TPT':
                    n31.append(name_s1[i])
                    h31.append(host_s1[i])

            #print('n11',n11,h11)
            #print('n21',n21,h21)
            #print('n31',n31,h31)
            

            #---------------prod data---------
            x11 = np.array(n11)
            u11=np.unique(x11)
            
            dict11= collections.defaultdict(list)
            for i in range(len(u11)):
                for j in range(len(n11)):
                    if n11[j]==u11[i]:
                        dict11[u11[i]].append(h11[j])

            p_name11=dict(dict11)
            #print('pn',p_name11)
                
                
            #------------------------non-prod data
           
            x21 = np.array(n21)
            u21=np.unique(x21)
            
            dict21= collections.defaultdict(list)
            for i in range(len(u21)):
               
                for j in range(len(n21)):
                   
                    if n21[j]==u21[i]:
                        dict21[u21[i]].append(h21[j])

            np_name21=dict(dict21)
            #print(np_name21)

            #------------------------TPT data
            x31 = np.array(n31)
            u31=np.unique(x31)
            #print(u31)
            
            dict31= collections.defaultdict(list)
            for i in range(len(u31)):
                for j in range(len(n31)):
                   
                    if n31[j]==u31[i]:
                        dict31[u31[i]].append(h31[j])

            t_name31=dict(dict31)
            #print(t_name31)             


            #------------------datanode table

            name_s3=list(df_bdn['Name'])
            host_s3=list(df_bdn['Hostname'])
            env_s3=list(df_bdn['Environment'])

            n13,n23,n33=[],[],[]
            h13,h23,h33=[],[],[]

            for i in range(len(name_s3)):
                if env_s3[i] == 'PROD':
                   n13.append(name_s3[i])
                   h13.append(host_s3[i])

                elif env_s3[i]=='NON-PROD':
                    n23.append(name_s3[i])
                    h23.append(host_s3[i])

                elif env_s3[i]=='TPT':
                    n33.append(name_s3[i])
                    h33.append(host_s3[i])

            #---------------prod data---------
            x13 = np.array(n13)
            u13=np.unique(x13)
            
            dict13= collections.defaultdict(list)
            for i in range(len(u13)):
                for j in range(len(n13)):
                    if n13[j]==u13[i]:
                        dict13[u13[i]].append(h13[j])

            p_name13=dict(dict13)
            #print('pn',p_name13)
                
                
            #------------------------non-prod data
           
            x23 = np.array(n23)
            u23=np.unique(x23)
            
            dict23= collections.defaultdict(list)
            for i in range(len(u23)):
               
                for j in range(len(n23)):
                   
                    if n23[j]==u23[i]:
                        dict23[u23[i]].append(h23[j])

            np_name23=dict(dict23)
            #print(np_name23)

            #------------------------TPT data
            x33 = np.array(n33)
            u33=np.unique(x33)
            #print(u31)
            
            dict33= collections.defaultdict(list)
            for i in range(len(u33)):
                for j in range(len(n33)):
                   
                    if n33[j]==u33[i]:
                        dict33[u33[i]].append(h33[j])

            t_name33=dict(dict33)
            #print(t_name33)             

            #-------------datanode count table design

            if datanode_count['PROD']==0:
                k=4
                sheet['N4']='NON-PROD'
                sheet['Q4']=datanode_count['NON-PROD']
                if datanode_count['NON-PROD']==0:
                    sheet['N4']='NON-PROD'
                    sheet['Q4']=0
                    sheet['N5']='TPT'
                    sheet['Q5']=datanode_count['TPT']
                    k=6
                    if datanode_count['TPT']==0:
                        sheet['N5']='TPT'
                        sheet['Q5']=0
                        sheet['N6']='Grand Total'
                        sheet['Q6']=datanode_count['TPT']+datanode_count['PROD']+datanode_count['NON-PROD']

                    else:
                        sheet['N5']='TPT'
                        sheet['Q5']=datanode_count['TPT']
                    
                        t_k3=list(t_name33.keys())

                        for i in t_k3:
                            sheet[str('O'+str(k))]=str(i)
                            t_v3=t_name33[i]
                            for j in range(len(t_v3)):
                                sheet[str('P'+str(k+1+j))]=str(t_v3[j])
                                sheet[str('Q'+str(k+1+j))]=1
                            sheet[str('N'+str(k+len(t_v3)+1))]=str(i +' SubTotal')
                            sheet[str('Q'+str(k+len(t_v3)+1))]=len(t_v3)
                            k=k+len(t_v3)+2

                        sheet[str('N'+str(k))]='Grand Total'
                        sheet[str('Q'+str(k))]=datanode_count['TPT']+datanode_count['PROD']+datanode_count['NON-PROD']


                else:
                    k=4
                    sheet[str('N'+str(k))]='NON-PROD'
                    sheet[str('Q'+str(k))]=datanode_count['NON-PROD']

                    np_k3=list(np_name23.keys())
                    
                    for i in np_k3:
                        sheet[str('O'+str(k))]=str(i)
                        np_v3=np_name23[i]
                        for j in range(len(np_v3)):
                            sheet[str('P'+str(k+1+j))]=str(np_v3[j])
                            sheet[str('Q'+str(k+1+j))]=1
                        sheet[str('N'+str(k+len(np_v3)+1))]=str(i +' SubTotal')
                        sheet[str('Q'+str(k+len(np_v3)+1))]=len(np_v3)
                        k=k+len(np_v3)+2

                    if datanode_count['TPT']==0:
                        sheet[str('N'+str(k))]='TPT'
                        sheet[str('Q'+str(k))]=0
                        sheet[str('N'+str(k+1))]='Grand Total'
                        sheet[str('Q'+str(k+1))]=datanode_count['TPT']+datanode_count['PROD']+datanode_count['NON-PROD']

                    else:
                        sheet[str('N'+str(k))]='TPT'
                        sheet[str('Q'+str(k))]=datanode_count['TPT']
                    
                        t_k3=list(t_name33.keys())

                        for i in t_k3:
                            sheet[str('O'+str(k))]=str(i)
                            t_v3=t_name33[i]
                            for j in range(len(t_v3)):
                                sheet[str('P'+str(k+1+j))]=str(t_v3[j])
                                sheet[str('Q'+str(k+1+j))]=1
                            sheet[str('N'+str(k+len(t_v3)+1))]=str(i +' SubTotal')
                            sheet[str('Q'+str(k+len(t_v3)+1))]=len(t_v1)
                            k=k+len(t_v3)+2

                        sheet[str('N'+str(k))]='Grand Total'
                        sheet[str('Q'+str(k))]=datanode_count['TPT']+datanode_count['PROD']+datanode_count['NON-PROD']

            else:
                p_k3=list(p_name13.keys())
                #print(p_k3)

                k=4
                for i in p_k3:
                    #print('i',i)
                    #print('O'+str(k))
                    sheet[str('O'+str(k))]=str(i)
                    p_v3=p_name13[i]
                    #print(p_v3)
                    for j in range(len(p_v3)):
                        #print('j',j)
                        #print('P'+str(k+1+j))
                        sheet[str('p'+str(k+1+j))]=str(p_v3[j])
                        #print(str('Q'+str(k+1+j)))
                        sheet[str('Q'+str(k+1+j))]=1
                    sheet[str('N'+str(k+len(p_v3)+1))]=str(i +' SubTotal')
                    sheet[str('Q'+str(k+len(p_v3)+1))]=len(p_v3)
                    k=k+len(p_v3)+2
                    #print(k)

                if datanode_count['NON-PROD']==0:
                    sheet[str('N'+str(k))]='NON-PROD'
                    sheet[str('Q'+str(k))]=0
                    sheet[str('N'+str(k+1))]='TPT'
                    sheet[str('Q'+str(k+1))]=datanode_count['TPT']
                    if datanode_count['TPT']==0:
                        sheet[str('N'+str(k))]='TPT'
                        sheet[str('Q'+str(k))]=0
                        print('N'+str(k+1))
                        sheet[str('N'+str(k+1))]='Grand Total'
                        sheet[str('Q'+str(k+1))]=datanode_count['TPT']+datanode_count['PROD']+datanode_count['NON-PROD']

                    else:
                        sheet[str('N'+str(k))]='TPT'
                        sheet[str('Q'+str(k))]=datanode_count['TPT']
                    
                        t_k3=list(t_name33.keys())

                        for i in t_k3:
                            sheet[str('O'+str(k))]=str(i)
                            t_v3=t_name33[i]
                            for j in range(len(t_v3)):
                                sheet[str('P'+str(k+1+j))]=str(t_v3[j])
                                sheet[str('Q'+str(k+1+j))]=1
                            sheet[str('N'+str(k+len(t_v3)+1))]=str(i +' SubTotal')
                            sheet[str('Q'+str(k+len(t_v3)+1))]=len(t_v3)
                            k=k+len(t_v3)+2

                        sheet[str('N'+str(k))]='Grand Total'
                        sheet[str('Q'+str(k))]=datanode_count['TPT']+datanode_count['PROD']+datanode_count['NON-PROD']


                else:
                    sheet[str('N'+str(k))]='NON-PROD'
                    sheet[str('Q'+str(k))]=datanode_count['NON-PROD']

                    np_k3=list(np_name23.keys())
                    
                    for i in np_k1:
                        sheet[str('O'+str(k))]=str(i)
                        np_v3=np_name23[i]
                        for j in range(len(np_v3)):
                            sheet[str('P'+str(k+1+j))]=str(np_v3[j])
                            sheet[str('Q'+str(k+1+j))]=1
                        sheet[str('N'+str(k+len(np_v3)+1))]=str(i +' SubTotal')
                        sheet[str('Q'+str(k+len(np_v3)+1))]=len(np_v3)
                        k=k+len(np_v3)+2

                    if datanode_count['TPT']==0:
                        sheet[str('N'+str(k))]='TPT'
                        sheet[str('Q'+str(k))]=0
                        #print('N'+str(k+1))
                        sheet[str('N'+str(k+1))]='Grand Total'
                        sheet[str('Q'+str(k+1))]=datanode_count['TPT']+datanode_count['PROD']+datanode_count['NON-PROD']

                    else:
                        sheet[str('N'+str(k))]='TPT'
                        sheet[str('Q'+str(k))]=datanode_count['TPT']
                    
                        t_k3=list(t_name33.keys())

                        for i in t_k3:
                            sheet[str('O'+str(k))]=str(i)
                            t_v3=t_name33[i]
                            for j in range(len(t_v3)):
                                sheet[str('P'+str(k+1+j))]=str(t_v3[j])
                                sheet[str('Q'+str(k+1+j))]=1
                            sheet[str('N'+str(k+len(t_v3)+1))]=str(i +' SubTotal')
                            sheet[str('Q'+str(k+len(t_v3)+1))]=len(t_v3)
                            k=k+len(t_v3)+2

                        sheet[str('N'+str(k))]='Grand Total'
                        sheet[str('Q'+str(k))]=datanode_count['TPT']+datanode_count['PROD']+datanode_count['NON-PROD']                  

            wb.save('workbook1.xlsx')

            #--------------------------Growth--------------------------------
            wb = load_workbook('workbook1.xlsx')
            ws1 = wb.create_sheet('sheet',0)
            ws1.title='Growth'
            wb.save(filename = 'workbook1.xlsx')
            wb = load_workbook('workbook1.xlsx')
            sheet = wb['Growth']
            
            #-----------------server
            s_rows=[['','SERVERS', '','',''],
                  ['MONTH-YEAR','PROD','TPT','DEV','TOTAL']]

            for i in s_rows:
                sheet.append(i)

            s_month=int(input("How many months details do you want to enter for servers?"))
            final_s=[]
            for i in range(s_month):
                s_month_year=input("Enter month-year for server details: ")
                print("-------- Server Details "+ s_month_year +" -----") 
                s_prod_count = int(input("Enter PROD count details for server: "))
                s_tpt_count = int(input("Enter TPT count details for server: "))
                s_dev_count = int(input("Enter DEV count details for server: "))
                s_total= s_prod_count + s_tpt_count + s_dev_count

                data=[s_month_year,s_prod_count,s_tpt_count,s_dev_count,s_total]
                #print(data)
                final_s.append(data)
            #print(len(final_s))
            #print(final_s)

            for i in final_s:
                sheet.append(i)

            for i in range(13):
                sheet.append(['','','','',''])

            Chart = LineChart()
            Chart.title = "Servers"
            Chart.style = 13
            Chart.y_axis.title = 'Environment count'
            Chart.x_axis.title = 'N(th/st/rd) Month'
            s_chart_data = Reference(worksheet=sheet,  
                             min_row=2,  
                             max_row=len(final_s)+2,  
                             min_col=1,max_col=4)  
              
            Chart.add_data(s_chart_data,from_rows=False, titles_from_data=True)

            # Style the lines
            s1 = Chart.series[0]
            s1.marker.symbol = "triangle"
            s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
            s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline

            s1.graphicalProperties.line.noFill = True

            s2 = Chart.series[1]
            s2.graphicalProperties.line.solidFill = "00AAAA"
            s2.graphicalProperties.line.dashStyle = "sysDot"
            s2.graphicalProperties.line.width = 100050 # width in EMUs

            s2 = Chart.series[2]
            s2.smooth = True # Make the line smooth

            s1 = Chart.series[3]
            s1.marker.symbol = "square"
            s1.marker.graphicalProperties.solidFill = "0000FF" # Marker filling
            s1.marker.graphicalProperties.line.solidFill = "0000FF" # Marker outline

            sheet.add_chart(Chart, "h2")

            wb.save('workbook1.xlsx')
                   
            #--------------------databases
            d_rows=[['','DATABASES', '','',''],
                  ['MONTH-YEAR','PROD','TPT','DEV','TOTAL']]

            for i in d_rows:
                sheet.append(i)

            d_month=int(input("How many months details do you want to enter for databases?"))
            final_d=[]
            for i in range(d_month):
                d_month_year=input("Enter month-year for database details: ")
                print("-------- Databases Details "+ d_month_year +" -----") 
                d_prod_count = int(input("Enter PROD count details for database: "))
                d_tpt_count = int(input("Enter TPT count details for database: "))
                d_dev_count = int(input("Enter DEV count details for database: "))
                d_total= d_prod_count + d_tpt_count + d_dev_count

                data=[d_month_year,d_prod_count,d_tpt_count,d_dev_count,d_total]
                final_d.append(data)
           # print(final_d)

            for i in final_d:
                sheet.append(i)
            for i in range(13):
                sheet.append(['','','','',''])

            Chart = LineChart()
            Chart.title = "Databases"
            Chart.style = 13
            Chart.y_axis.title = 'Environment count'
            Chart.x_axis.title = 'N(th/st/rd) Month'
            d_chart_data = Reference(worksheet=sheet,  
                             min_row=len(final_s)+17,  
                             max_row=len(final_d)+len(final_s)+17,  
                             min_col=1,max_col=4)  
              
            Chart.add_data(d_chart_data,from_rows=False, titles_from_data=True)
            # Style the lines
            s1 = Chart.series[0]
            s1.marker.symbol = "triangle"
            s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
            s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline

            s1.graphicalProperties.line.noFill = True

            s2 = Chart.series[1]
            s2.graphicalProperties.line.solidFill = "00AAAA"
            s2.graphicalProperties.line.dashStyle = "sysDot"
            s2.graphicalProperties.line.width = 100050 # width in EMUs

            s2 = Chart.series[2]
            s2.smooth = True # Make the line smooth

            s1 = Chart.series[3]
            s1.marker.symbol = "square"
            s1.marker.graphicalProperties.solidFill = "0000FF" # Marker filling
            s1.marker.graphicalProperties.line.solidFill = "0000FF" # Marker outline
            sheet.add_chart(Chart, "h20")

            wb.save('workbook1.xlsx')
        
            #--------------------cluster
            c_rows=[['','CLUSTERS', '','',''],
                  ['MONTH-YEAR','PROD','TPT','DEV','TOTAL','']]

            for i in c_rows:
                sheet.append(i)

            c_month=int(input("How many months details do you want to enter for clusters?"))
            final_c=[]
            for i in range(c_month):
                c_month_year=input("Enter month-year for cluster details: ")
                print("-------- Cluster Details "+ c_month_year +" -----") 
                c_prod_count = int(input("Enter PROD count details for cluster: "))
                c_tpt_count = int(input("Enter TPT count details for cluster: "))
                c_dev_count = int(input("Enter DEV count details for cluster: "))
                c_total= c_prod_count + c_tpt_count + c_dev_count

                data=[c_month_year,c_prod_count,c_tpt_count,c_dev_count,c_total]
                final_c.append(data)
           # print(final_c)

            for i in final_c:
                sheet.append(i)

            for i in range(13):
                sheet.append(['','','','',''])
            
            Chart = LineChart()
            Chart.title = "Clusters"
            Chart.style = 13
            Chart.y_axis.title = 'Environment count'
            Chart.x_axis.title = 'N(th/st/rd) Month'
            c_chart_data = Reference(worksheet=sheet,  
                             min_row=len(final_d)+len(final_s)+17+15,  
                             max_row=len(final_c)+len(final_d)+len(final_s)+32,  
                             min_col=1,max_col=4)  
              
            Chart.add_data(c_chart_data,from_rows=False, titles_from_data=True)
            # Style the lines
            s1 = Chart.series[0]
            s1.marker.symbol = "triangle"
            s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
            s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline

            s1.graphicalProperties.line.noFill = True

            s2 = Chart.series[1]
            s2.graphicalProperties.line.solidFill = "00AAAA"
            s2.graphicalProperties.line.dashStyle = "sysDot"
            s2.graphicalProperties.line.width = 100050 # width in EMUs

            s2 = Chart.series[2]
            s2.smooth = True # Make the line smooth

            s1 = Chart.series[3]
            s1.marker.symbol = "square"
            s1.marker.graphicalProperties.solidFill = "0000FF" # Marker filling
            s1.marker.graphicalProperties.line.solidFill = "0000FF" # Marker outline
            sheet.add_chart(Chart, "h39")

            wb.save('workbook1.xlsx')
            
            #--------------------datanodes
            dn_rows=[['','DATANODES', '','',''],
                  ['MONTH-YEAR','PROD','TPT','DEV','TOTAL','']]

            for i in dn_rows:
                sheet.append(i)

            dn_month=int(input("How many months details do you want to enter for datanodes?"))
            final_dn=[]
            for i in range(dn_month):
                dn_month_year=input("Enter month-year for datanode details: ")
                print("-------- Datanode Details "+ dn_month_year +" -----") 
                dn_prod_count = int(input("Enter PROD count details for datanode: "))
                dn_tpt_count = int(input("Enter TPT count details for datanode: "))
                dn_dev_count = int(input("Enter DEV count details for datanode: "))
                dn_total= dn_prod_count + dn_tpt_count + dn_dev_count

                data=[dn_month_year,dn_prod_count,dn_tpt_count,dn_dev_count,dn_total]
                final_dn.append(data)
            #print(final_dn)

            for i in final_dn:
                sheet.append(i)

            for i in range(13):
                sheet.append(['','','','',''])

            Chart = LineChart()
            Chart.title = "Datanodes"
            Chart.style = 13
            Chart.y_axis.title = 'Environment count'
            Chart.x_axis.title = 'N(th/st/rd) Month'
            dn_chart_data = Reference(worksheet=sheet,  
                             min_row=len(final_c)+len(final_d)+len(final_s)+17+15+15,  
                             max_row=len(final_dn)+len(final_c)+len(final_d)+len(final_s)+47,  
                             min_col=1,max_col=4)  
              
            Chart.add_data(dn_chart_data,from_rows=False, titles_from_data=True)
            # Style the lines
            s1 = Chart.series[0]
            s1.marker.symbol = "triangle"
            s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
            s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline

            s1.graphicalProperties.line.noFill = True

            s2 = Chart.series[1]
            s2.graphicalProperties.line.solidFill = "00AAAA"
            s2.graphicalProperties.line.dashStyle = "sysDot"
            s2.graphicalProperties.line.width = 100050 # width in EMUs

            s2 = Chart.series[2]
            s2.smooth = True # Make the line smooth

            s1 = Chart.series[3]
            s1.marker.symbol = "square"
            s1.marker.graphicalProperties.solidFill = "0000FF" # Marker filling
            s1.marker.graphicalProperties.line.solidFill = "0000FF" # Marker outline
            sheet.add_chart(Chart, "h55")
            wb.save('workbook1.xlsx')

            #-------------------------Summary--------------------------------
            
            wb = load_workbook('workbook1.xlsx')
            ws1 = wb.create_sheet('sheet',0)
            ws1.title='Summary'
            wb.save(filename = 'workbook1.xlsx')
            wb = load_workbook('workbook1.xlsx')
            sheet = wb['Summary']
            
            
            rows=[['ENV','Cluster Count', 'Server Count','Database Count'],
                  ['PROD',cluster_count['PROD'],Server_count['PROD'],database_count['PROD']],
                  ['TPT',cluster_count['TPT'],Server_count['TPT'],database_count['TPT']],
                  ['NON-PROD',cluster_count['NON-PROD'],Server_count['NON-PROD'],database_count['NON-PROD']]
                  ]

            for i in rows:
                sheet.append(i)

            c_t=int(cluster_count['PROD'])+int(cluster_count['TPT'])+int(cluster_count['NON-PROD'])
            s_t=int(Server_count['PROD'])+int(Server_count['TPT'])+int(Server_count['NON-PROD'])
            d_t=int(database_count['PROD'])+int(database_count['TPT'])+int(database_count['NON-PROD'])

            sum_count=['Total',c_t,s_t,d_t]
            sheet.append(sum_count)
            sheet.auto_filter.ref = "A1:D1"

            wb.save('workbook1.xlsx')
        
        else:
            file=open('workbook1.xlsx','w')
            file.close()
            print("please run script again")

    except Exception as e:
     print(e)

csvFileRead()
